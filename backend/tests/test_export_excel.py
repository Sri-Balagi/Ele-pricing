import pytest
from io import BytesIO
from openpyxl import load_workbook
from app.export.excel_exporter import ExcelExporter
from app.models.domain import ExportContext
from app.core.constants import ExportFormat

def test_excel_export_and_integrity(sample_configuration):
    exporter = ExcelExporter()
    context = ExportContext(
        configuration=sample_configuration,
        correlation_id="test-123",
        execution_timestamp="2026-07-08T00:00:00Z",
        export_format=ExportFormat.EXCEL
    )
    
    report = exporter.export(context)
    assert report.success
    assert report.content is not None
    assert report.file_size > 0
    assert report.export_format == ExportFormat.EXCEL
    
    # Verify integrity of workbook
    wb = load_workbook(BytesIO(report.content))
    expected_sheets = ["Quote Summary", "Bill Of Materials", "Pricing Breakdown", "Configuration Summary"]
    assert wb.sheetnames == expected_sheets
    
    # Check headers
    qs = wb["Quote Summary"]
    assert qs["A1"].value == "Field"
    assert qs["B1"].value == "Value"
