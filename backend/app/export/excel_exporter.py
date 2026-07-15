import hashlib
import time
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.core.constants import ExportFormat
from app.export.base_exporter import BaseExporter
from app.models.domain import ExportContext, ExportReport


class ExcelExporter(BaseExporter):
    """
    Exports the Configuration to a formatted Excel Workbook.
    Includes Freeze Panes, Auto Width, Auto Filter, and Bold Headers.
    """

    def export(self, context: ExportContext) -> ExportReport:
        t0 = time.perf_counter()
        config = context.configuration

        wb = Workbook()

        # Sheet 1: Quote Summary
        ws1 = wb.active
        ws1.title = "Quote Summary"
        self._write_headers(ws1, ["Field", "Value"])

        quote = config.quote_metadata
        ws1.append(["Quote Number", quote.quote_number if quote else "DRAFT"])
        ws1.append(["Revision", quote.quote_version if quote else 1])
        ws1.append(["Valid Until", quote.valid_until if quote else "N/A"])
        ws1.append(["Customer Reference", config.customer_reference or "N/A"])

        price = config.pricing_summary
        if price:
            ws1.append(
                ["Total After Tax", f"{price.currency_symbol}{price.total_after_tax}"]
            )

        self._apply_formatting(ws1)

        # Sheet 2: Bill Of Materials
        ws2 = wb.create_sheet(title="Bill Of Materials")
        bom_headers = [
            "Component ID",
            "Component Name",
            "Qty",
            "Unit Cost"
        ]
        self._write_headers(ws2, bom_headers)

        if config.bill_of_materials:
            # Sort Base first, then features
            base_items = [i for i in config.bill_of_materials.items if getattr(i.origin_type, "value", str(i.origin_type)) == "BASE"]
            feature_items = [i for i in config.bill_of_materials.items if getattr(i.origin_type, "value", str(i.origin_type)) != "BASE"]
            
            for item in base_items + feature_items:
                comp_name = item.component_id
                if context.catalogue:
                    comp = next((c for c in context.catalogue.components if c.id == item.component_id), None)
                    if comp:
                        comp_name = comp.name
                
                cost_val = item.unit_cost if item.unit_cost is not None else 0.0
                ws2.append(
                    [
                        item.component_id,
                        comp_name,
                        item.quantity,
                        f"${cost_val:.2f}"
                    ]
                )
        self._apply_formatting(ws2)

        # Sheet 3: Pricing Breakdown
        ws3 = wb.create_sheet(title="Pricing Breakdown")
        price_headers = ["Component/Feature", "Unit Cost", "Quantity", "Total"]
        self._write_headers(ws3, price_headers)

        if price:
            ws3.append(["Base Cost", "", "", f"{price.currency_symbol}{price.category_cost:.2f}"])
            ws3.append(["Feature Cost", "", "", f"{price.currency_symbol}{price.feature_cost:.2f}"])
            if price.floor_coverage_cost > 0:
                ws3.append(["Additional Floor Coverage", "", "", f"{price.currency_symbol}{price.floor_coverage_cost:.2f}"])
            ws3.append(["Subtotal before tax", "", "", f"{price.currency_symbol}{price.subtotal_before_tax:.2f}"])
            ws3.append(["Tax amount", "", "", f"{price.currency_symbol}{price.tax_amount:.2f}"])
            ws3.append(["Grand total", "", "", f"{price.currency_symbol}{price.total_after_tax:.2f}"])
        self._apply_formatting(ws3)

        # Sheet 4: Configuration Summary
        ws4 = wb.create_sheet(title="Configuration Summary")
        self._write_headers(ws4, ["Feature", "Option Name", "Option ID"])
        for opt_id in config.selected_feature_options:
            feat_name = "Unknown"
            opt_name = "Unknown"
            if context.catalogue:
                opt = next((o for o in context.catalogue.feature_options if o.id == opt_id), None)
                if opt:
                    opt_name = opt.display_name
                    feat = next((f for f in context.catalogue.features if f.id == opt.feature_id), None)
                    if feat:
                        feat_name = feat.name
            ws4.append([feat_name, opt_name, opt_id])
        self._apply_formatting(ws4)

        # Save to bytes
        buffer = BytesIO()
        wb.save(buffer)
        content_bytes = buffer.getvalue()

        checksum = hashlib.sha256(content_bytes).hexdigest()
        file_size = len(content_bytes)
        generation_time_ms = (time.perf_counter() - t0) * 1000

        return ExportReport(
            export_format=ExportFormat.EXCEL,
            filename=f"{config.configuration_id}.xlsx",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            checksum=checksum,
            file_size=file_size,
            generation_time_ms=generation_time_ms,
            success=True,
            content=content_bytes,
        )

    def _write_headers(self, ws, headers):
        ws.append(headers)

    def _apply_formatting(self, ws):
        """Applies Bold Headers, Freeze Panes, Auto Filter, and Auto Width."""
        # Bold headers
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Freeze top row
        ws.freeze_panes = "A2"

        # Auto Filter
        if ws.max_row > 1 and ws.max_column > 0:
            last_col_letter = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

        # Auto Width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width
